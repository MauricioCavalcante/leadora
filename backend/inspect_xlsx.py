import openpyxl

wb = openpyxl.load_workbook('../Controle-Mari.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet.")
        continue

    # first 5 rows
    for i, r in enumerate(rows[:5]):
        print(f"Row {i}: {r}")
