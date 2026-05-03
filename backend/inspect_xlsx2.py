import openpyxl
wb = openpyxl.load_workbook('../Controle-Mari.xlsx', data_only=True)
sheet = wb[wb.sheetnames[0]]
print("Sheet 1 Name:", wb.sheetnames[0])
print("Header Row:", sheet[1])
for cell in sheet[1]:
    print(cell.value)
